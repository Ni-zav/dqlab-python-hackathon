from __future__ import annotations

import math
from itertools import combinations
from pathlib import Path

import pandas as pd


MIN_SUPPORT = 0.01
MIN_LIFT = 2.0
MIN_CONSECUTIVE_DAYS = 12
REQUIRED_COLUMNS = [
    "nomor_struk",
    "tgl_transaksi",
    "kode_produk",
    "nama_produk",
    "jumlah_terjual",
    "harga",
    "total_nilai",
]


def find_input_file() -> Path:
    names = [
        "data_penjualan.xlsx",
        "sales_transaction.xlsx",
        "sales_transactions.xlsx",
        "sales_transaction.csv",
        "sales_transactions.csv",
    ]
    dirs = [
        Path("."),
        Path("tasks"),
        Path("../tasks"),
        Path("../../tasks"),
        Path("../../../tasks"),
    ]
    for directory in dirs:
        for name in names:
            path = directory / name
            if path.exists():
                return path

    required = set(REQUIRED_COLUMNS)
    for pattern in ("*.xlsx", "*.xls", "*.csv"):
        for path in Path(".").glob(pattern):
            if path.stem.lower().startswith("retail_insight"):
                continue
            try:
                columns = (
                    pd.read_excel(path, nrows=0).columns
                    if path.suffix.lower() in {".xlsx", ".xls"}
                    else pd.read_csv(path, nrows=0).columns
                )
            except Exception:
                continue
            if required.issubset(set(columns)):
                return path

    raise FileNotFoundError("Dataset transaksi tidak ditemukan.")


def read_sales(path: Path) -> pd.DataFrame:
    if path.suffix.lower() in {".xlsx", ".xls"}:
        df = read_excel_fast(path)
    else:
        df = pd.read_csv(path, usecols=REQUIRED_COLUMNS)
    df["tgl_transaksi"] = pd.to_datetime(df["tgl_transaksi"], dayfirst=True)
    return df


def read_excel_fast(path: Path) -> pd.DataFrame:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    header = list(next(rows))
    indexes = [header.index(column) for column in REQUIRED_COLUMNS]
    data = [tuple(row[index] for index in indexes) for row in rows]
    workbook.close()
    return pd.DataFrame(data, columns=REQUIRED_COLUMNS)


def build_daily_sales(df: pd.DataFrame) -> pd.DataFrame:
    daily = (
        df.groupby(["kode_produk", "nama_produk", "tgl_transaksi"], sort=False, as_index=False)[
            "total_nilai"
        ]
        .sum()
        .sort_values(["kode_produk", "tgl_transaksi"], kind="mergesort")
    )
    daily["MA"] = daily.groupby("kode_produk", sort=False)["total_nilai"].transform(
        lambda values: values.rolling(3, min_periods=3).mean()
    )
    daily["Normalized"] = daily.groupby("kode_produk", sort=False)["MA"].transform(
        normalize_base_100
    )
    return daily


def normalize_base_100(values: pd.Series) -> pd.Series:
    valid = values.dropna()
    if valid.empty or valid.iloc[0] == 0:
        return values * pd.NA
    return values / valid.iloc[0] * 100


def find_rising_stars(df: pd.DataFrame, daily: pd.DataFrame) -> pd.DataFrame:
    total_sales = df.groupby("kode_produk", sort=False)["total_nilai"].sum()
    rows = []

    for (code, name), group in daily.groupby(["kode_produk", "nama_produk"], sort=False):
        ma = group["MA"].tolist()
        best_growth = None
        run_start = None
        run_length = 0

        for idx in range(1, len(ma)):
            today = ma[idx]
            yesterday = ma[idx - 1]
            increasing = pd.notna(today) and pd.notna(yesterday) and today > yesterday
            if increasing:
                if run_length == 0:
                    run_start = idx
                run_length += 1
                continue

            best_growth = update_best_growth(best_growth, ma, run_start, idx - 1, run_length)
            run_start = None
            run_length = 0

        best_growth = update_best_growth(best_growth, ma, run_start, len(ma) - 1, run_length)
        if best_growth is not None:
            rows.append(
                {
                    "Kode Produk": code,
                    "Nama Produk": name,
                    "Growth %": round(best_growth, 2),
                    "Total Penjualan": int(total_sales.loc[code]),
                }
            )

    columns = ["Kode Produk", "Nama Produk", "Growth %", "Total Penjualan"]
    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(rows, columns=columns).sort_values(
        ["Growth %", "Total Penjualan"], ascending=[False, False], ignore_index=True
    )


def update_best_growth(best_growth, ma, start_idx, end_idx, run_length):
    if start_idx is None or run_length < MIN_CONSECUTIVE_DAYS:
        return best_growth
    start_ma = ma[start_idx]
    end_ma = ma[end_idx]
    if pd.isna(start_ma) or pd.isna(end_ma) or start_ma == 0:
        return best_growth
    growth = (end_ma - start_ma) / start_ma * 100
    return growth if best_growth is None or growth > best_growth else best_growth


def build_packaging(df: pd.DataFrame, rising: pd.DataFrame) -> pd.DataFrame:
    product_ref = df.drop_duplicates("kode_produk")
    product_names = product_ref.set_index("kode_produk")["nama_produk"].to_dict()
    product_order = {code: idx for idx, code in enumerate(product_ref["kode_produk"])}
    rising_codes = set(rising["Kode Produk"])
    rising_order = {code: idx for idx, code in enumerate(rising["Kode Produk"])}

    rules = fast_apriori_rules(df, rising_codes)
    rows = []
    for rule in rules:
        antecedents = sort_bundle(rule["antecedents"], rising_order, product_order)
        consequents = sort_bundle(rule["consequents"], rising_order, product_order)
        rows.append(
            {
                "Jika Membeli": ", ".join(product_names[code] for code in antecedents),
                "Maka Membeli": ", ".join(product_names[code] for code in consequents),
                "Jumlah Invoice": rule["count"],
                "Support": round(rule["support"], 2),
                "Confidence": round(rule["confidence"], 2),
                "Lift": round(rule["lift"], 2),
                "_support": rule["support"],
                "_confidence": rule["confidence"],
                "_lift": rule["lift"],
            }
        )

    columns = [
        "Jika Membeli",
        "Maka Membeli",
        "Jumlah Invoice",
        "Support",
        "Confidence",
        "Lift",
    ]
    if not rows:
        return pd.DataFrame(columns=columns)
    return (
        pd.DataFrame(rows)
        .sort_values(["_lift", "_support", "_confidence"], ascending=[False, False, False])
        .loc[:, columns]
        .reset_index(drop=True)
    )


def sort_bundle(codes, rising_order, product_order):
    return sorted(
        codes,
        key=lambda code: (0 if code in rising_order else 1, rising_order.get(code, product_order[code])),
    )


def fast_apriori_rules(df: pd.DataFrame, rising_codes: set[str]) -> list[dict]:
    code_order = list(dict.fromkeys(df["kode_produk"]))
    code_to_id = {code: idx for idx, code in enumerate(code_order)}
    id_to_code = {idx: code for code, idx in code_to_id.items()}
    rising_ids = {code_to_id[code] for code in rising_codes if code in code_to_id}

    invoice_map: dict[str, set[int]] = {}
    for invoice, code in zip(df["nomor_struk"], df["kode_produk"]):
        invoice_map.setdefault(invoice, set()).add(code_to_id[code])
    transactions = [tuple(sorted(items)) for items in invoice_map.values()]

    invoice_count = len(transactions)
    min_count = math.ceil(MIN_SUPPORT * invoice_count)
    support_count = count_frequent_itemsets(transactions, min_count)

    rules = []
    for itemset, itemset_count in support_count.items():
        if len(itemset) < 2:
            continue
        for size in range(1, len(itemset)):
            for antecedent in combinations(itemset, size):
                consequent = tuple(item for item in itemset if item not in antecedent)
                if not (set(antecedent) & rising_ids or set(consequent) & rising_ids):
                    continue
                confidence = itemset_count / support_count[antecedent]
                lift = confidence / (support_count[consequent] / invoice_count)
                if lift >= MIN_LIFT:
                    rules.append(
                        {
                            "antecedents": frozenset(id_to_code[item] for item in antecedent),
                            "consequents": frozenset(id_to_code[item] for item in consequent),
                            "count": itemset_count,
                            "support": itemset_count / invoice_count,
                            "confidence": confidence,
                            "lift": lift,
                        }
                    )
    return rules


def count_frequent_itemsets(transactions: list[tuple[int, ...]], min_count: int) -> dict[tuple[int, ...], int]:
    counts: dict[tuple[int, ...], int] = {}
    for transaction in transactions:
        for item in transaction:
            key = (item,)
            counts[key] = counts.get(key, 0) + 1

    current = {itemset: count for itemset, count in counts.items() if count >= min_count}
    support_count = dict(current)
    size = 2

    while current:
        candidates = generate_candidates(current)
        if not candidates:
            break

        counts = {}
        for transaction in transactions:
            if len(transaction) < size:
                continue
            for candidate in combinations(transaction, size):
                if candidate in candidates:
                    counts[candidate] = counts.get(candidate, 0) + 1

        current = {itemset: count for itemset, count in counts.items() if count >= min_count}
        support_count.update(current)
        size += 1

    return support_count


def generate_candidates(previous: dict[tuple[int, ...], int]) -> set[tuple[int, ...]]:
    previous_itemsets = sorted(previous)
    previous_set = set(previous_itemsets)
    candidates = set()

    for left_idx, left in enumerate(previous_itemsets):
        for right in previous_itemsets[left_idx + 1 :]:
            if left[:-1] != right[:-1]:
                break
            candidate = left + (right[-1],)
            if all(candidate[:idx] + candidate[idx + 1 :] in previous_set for idx in range(len(candidate))):
                candidates.add(candidate)

    return candidates


def write_excel(rising: pd.DataFrame, packaging: pd.DataFrame) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    write_sheet(workbook.active, "Rising Star", rising)
    write_sheet(workbook.create_sheet("Potential Packaging"), "Potential Packaging", packaging)
    workbook.save("retail_insight.xlsx")


def write_sheet(worksheet, title: str, frame: pd.DataFrame) -> None:
    worksheet.title = title
    worksheet.append(list(frame.columns))
    for row in frame.itertuples(index=False, name=None):
        worksheet.append(list(row))


def plot_outputs(df: pd.DataFrame, daily: pd.DataFrame, rising: pd.DataFrame) -> None:
    if rising.empty:
        return

    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    rising_codes = rising["Kode Produk"].tolist()
    top3_codes = (
        df.groupby(["kode_produk", "nama_produk"], sort=False)["total_nilai"]
        .sum()
        .reset_index()
        .sort_values("total_nilai", ascending=False)
        .head(3)["kode_produk"]
        .tolist()
    )
    top3 = daily[daily["kode_produk"].isin(top3_codes)]
    rising_daily = daily[daily["kode_produk"].isin(rising_codes)]

    palette = ["#FFD700", "#C0C0C0", "#CD7F32", "#2ecc71", "#3498db", "#9b59b6", "#e74c3c", "#34495e"]
    sorted_rising = rising.sort_values("Growth %", ascending=False)
    colors = {
        row["Kode Produk"]: palette[idx] if idx < len(palette) else "#95a5a6"
        for idx, (_, row) in enumerate(sorted_rising.iterrows())
    }
    ranks = {row["Kode Produk"]: idx + 1 for idx, (_, row) in enumerate(sorted_rising.iterrows())}

    draw_plot(
        plt,
        top3,
        rising_daily,
        colors,
        ranks,
        y_column="Normalized",
        title="ANALISIS PERTUMBUHAN RELATIF PRODUK RISING STAR\n(Dengan Benchmark Top 3 Total Penjualan)",
        ylabel="Indeks Pertumbuhan (Base 100)",
        filename="rising_star_index.png",
        baseline=True,
    )
    draw_plot(
        plt,
        top3,
        rising_daily,
        colors,
        ranks,
        y_column="total_nilai",
        title="ANALISIS NILAI PENJUALAN PRODUK RISING STAR\n(Nilai Penjualan Asli)",
        ylabel="Total Nilai Penjualan",
        filename="rising_star_actual.png",
        baseline=False,
    )


def draw_plot(plt, top3, rising_daily, colors, ranks, y_column, title, ylabel, filename, baseline):
    fig = plt.figure(figsize=(15, 8), dpi=100)
    ax = fig.add_subplot(111)
    grey_colors = ["#B0B0B0", "#909090", "#707070"]

    for idx, (code, group) in enumerate(top3.groupby("kode_produk")):
        ax.plot(
            group["tgl_transaksi"],
            group[y_column],
            linestyle="--",
            linewidth=2,
            marker="o",
            markersize=3,
            color=grey_colors[idx] if idx < len(grey_colors) else "#808080",
            alpha=0.7,
            label=f"Top Sales: {group['nama_produk'].iloc[0]}",
        )

    for code, group in rising_daily.groupby("kode_produk"):
        ax.plot(
            group["tgl_transaksi"],
            group[y_column],
            marker="o",
            markersize=4,
            linewidth=2.5,
            color=colors.get(code, "#95a5a6"),
            label=f"Rank {ranks[code]}: {group['nama_produk'].iloc[0]}",
        )

    ax.set_title(title, fontdict={"weight": "bold", "size": 16}, pad=20)
    ax.set_xlabel("Periode Tanggal", fontdict={"size": 12}, labelpad=10)
    ax.set_ylabel(ylabel, fontdict={"size": 12}, labelpad=10)
    ax.grid(True, linestyle="--", linewidth=0.5, alpha=0.5)
    if baseline:
        ax.axhline(y=100, color="black", linestyle="-", linewidth=1, alpha=0.5)
    plt.xticks(rotation=45, ha="right", fontsize=10)
    plt.yticks(fontsize=10)
    sort_legend(ax)
    plt.tight_layout()
    plt.savefig(filename, bbox_inches="tight")
    plt.close(fig)


def sort_legend(ax) -> None:
    handles, labels = ax.get_legend_handles_labels()
    top_sales = []
    rising = []
    for handle, label in zip(handles, labels):
        if label.startswith("Top Sales"):
            top_sales.append((handle, label))
        else:
            rising.append((handle, label))
    rising.sort(key=lambda item: int(item[1].split(":")[0].split()[1]))
    final = top_sales + rising
    ax.legend(
        [item[0] for item in final],
        [item[1] for item in final],
        title="Kategori Produk",
        title_fontsize=12,
        fontsize=10,
        bbox_to_anchor=(1.02, 1),
        loc="upper left",
        borderaxespad=0,
        frameon=True,
        shadow=True,
    )


def main() -> None:
    df = read_sales(find_input_file())
    daily = build_daily_sales(df)
    rising = find_rising_stars(df, daily)
    packaging = build_packaging(df, rising)
    write_excel(rising, packaging)
    plot_outputs(df, daily, rising)
    print("retail_insight.xlsx, rising_star_index.png, rising_star_actual.png dibuat.")


if __name__ == "__main__":
    main()
